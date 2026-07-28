"""Lectura de los diccionarios de códigos del REM.

Los archivos de datos del REM traen columnas **genéricas**: `Col01` a `Col38`, sin nombre.
Lo que cuenta cada una depende del `CodigoPrestacion` de la fila. El significado vive en un
diccionario aparte, uno por año, que **no es una tabla de consulta sino una réplica en Excel
del formulario en papel**: celdas combinadas, encabezados de tres niveles y una hoja por
sección. Este módulo lo traduce a algo que un programa pueda usar.

La estructura de una hoja, verificada sobre los diccionarios de 2010 a 2025:

    columna A       CodigoPrestacion (P6221600)
    columna B       grupo de conceptos (TRASTORNOS DEL HUMOR (AFECTIVOS))
    columna C       concepto específico (DEPRESIÓN LEVE)
    columna D+      COL01, COL02, ... que apuntan a las columnas del archivo de datos

    fila 10         sección del encabezado (T O T A L / GRUPO DE EDAD)
    fila 11         grupo etario, combinado sobre las dos columnas de sexo
    fila 12         sexo (Ambos sexos / Hombres / Mujeres)
    fila 13+        los datos: una fila por concepto

Las filas de encabezado se buscan por contenido y no por número: la fila 13 es donde
empieza el detalle en 2023, pero no en todos los años, y anclar en un número es la forma
más común de leer mal un formulario que se reordena.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

#: Un `CodigoPrestacion` del REM: la letra de la serie y seis a ocho dígitos.
PATRON_CODIGO = re.compile(r"^[ABDP]\d{6,8}$")
#: El nombre con que el diccionario nombra una columna del archivo de datos.
PATRON_COLUMNA = re.compile(r"^COL\s*(\d{1,3})$", re.I)

SEXOS = {
    "ambos sexos": "ambos",
    "ambos": "ambos",
    "hombres": "hombres",
    "mujeres": "mujeres",
    "hombre": "hombres",
    "mujer": "mujeres",
}


@dataclass(frozen=True)
class ColumnaRem:
    """Una columna del archivo de datos, con la dimensión que representa."""

    nombre: str          # "COL04"
    numero: int          # 4
    grupo_edad: str      # "0 a 4 años"; vacío cuando es la columna de total
    sexo: str            # "ambos" | "hombres" | "mujeres" | ""

    @property
    def es_total(self) -> bool:
        return not self.grupo_edad


@dataclass(frozen=True)
class ConceptoRem:
    """Una fila del formulario: qué se cuenta bajo un `CodigoPrestacion`."""

    codigo: str
    grupo: str
    concepto: str
    columnas: tuple[str, ...]   # nombres de las COLNN que usa esta fila

    @property
    def etiqueta(self) -> str:
        """Nombre legible, sin repetir el grupo cuando el concepto ya lo dice."""
        if not self.concepto or self.concepto == self.grupo:
            return self.grupo
        return f"{self.grupo} · {self.concepto}" if self.grupo else self.concepto


def _texto(v: object) -> str:
    return " ".join(str(v).split()) if v is not None else ""


def _celdas_xlsx(ruta: Path, hoja: str):
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(ruta, data_only=True)
    if hoja not in wb.sheetnames:
        return None
    ws = wb[hoja]
    return (lambda r, c: ws.cell(r, c).value), ws.max_row, ws.max_column


def _celdas_xls(ruta: Path, hoja: str):
    import xlrd  # noqa: PLC0415

    wb = xlrd.open_workbook(str(ruta), on_demand=True)
    if hoja not in wb.sheet_names():
        return None
    sh = wb.sheet_by_name(hoja)

    def celda(r: int, c: int):
        if r > sh.nrows or c > sh.ncols or r < 1 or c < 1:
            return None
        return sh.cell_value(r - 1, c - 1)

    return celda, sh.nrows, sh.ncols


def _abrir(ruta: Path, hoja: str):
    ruta = Path(ruta)
    lector = _celdas_xls if ruta.suffix.lower() == ".xls" else _celdas_xlsx
    return lector(ruta, hoja)


def _fila_de_columnas(celda, nfilas: int, ncols: int) -> int | None:
    """Encuentra la fila que nombra las COLNN. Se busca, no se asume."""
    for r in range(1, min(nfilas, 40) + 1):
        encontradas = sum(
            1 for c in range(1, min(ncols, 60) + 1) if PATRON_COLUMNA.match(_texto(celda(r, c)))
        )
        if encontradas >= 3:
            return r
    return None


def leer_columnas(ruta: Path, hoja: str = "P6") -> list[ColumnaRem]:
    """Devuelve las columnas del archivo de datos con su grupo etario y sexo.

    El grupo etario está en celdas combinadas que cubren las dos columnas de sexo, así que
    solo aparece en la primera; se arrastra hacia la derecha. Sin ese arrastre, la mitad de
    las columnas quedaría sin grupo y se leería como si fuera un total.
    """
    abierto = _abrir(ruta, hoja)
    if abierto is None:
        return []
    celda, nfilas, ncols = abierto
    fila_col = _fila_de_columnas(celda, nfilas, ncols)
    if fila_col is None:
        return []

    salida: list[ColumnaRem] = []
    edad_actual = ""
    for c in range(1, ncols + 1):
        m = PATRON_COLUMNA.match(_texto(celda(fila_col, c)))
        if not m:
            continue
        edad = _texto(celda(fila_col - 2, c))
        if edad:
            edad_actual = edad
        sexo_crudo = _texto(celda(fila_col - 1, c)).lower()
        sexo = SEXOS.get(sexo_crudo, "")
        # La columna de total no lleva grupo etario; el encabezado dice «T O T A L»
        # con espacios, que no calza con ningún grupo.
        es_total = "total" in edad_actual.lower().replace(" ", "")
        salida.append(
            ColumnaRem(
                nombre=f"COL{int(m.group(1)):02d}",
                numero=int(m.group(1)),
                grupo_edad="" if es_total else edad_actual,
                sexo=sexo,
            )
        )
    return salida


def leer_conceptos(ruta: Path, hoja: str = "P6") -> list[ConceptoRem]:
    """Devuelve un concepto por fila del formulario que tenga `CodigoPrestacion`.

    El grupo (columna B) se arrastra hacia abajo: el formulario lo escribe una sola vez y
    las filas siguientes lo heredan, igual que en el papel.
    """
    abierto = _abrir(ruta, hoja)
    if abierto is None:
        return []
    celda, nfilas, ncols = abierto

    salida: list[ConceptoRem] = []
    grupo_actual = ""
    for r in range(1, nfilas + 1):
        codigo = _texto(celda(r, 1)).upper()
        b, c_ = _texto(celda(r, 2)), _texto(celda(r, 3))
        if not PATRON_CODIGO.match(codigo):
            # Una fila con texto en B y sin código es un encabezado de grupo.
            if b and not codigo:
                grupo_actual = b
            continue
        if b:
            grupo_actual = b
        columnas = tuple(
            f"COL{int(m.group(1)):02d}"
            for col in range(4, ncols + 1)
            if (m := PATRON_COLUMNA.match(_texto(celda(r, col))))
        )
        salida.append(
            ConceptoRem(codigo=codigo, grupo=grupo_actual, concepto=c_, columnas=columnas)
        )
    return salida
