"""Tests del lector de diccionarios del REM.

El fixture se construye acá y no se guarda como archivo porque lo que hay que reproducir
son las **celdas combinadas** del formulario original: el grupo etario se escribe una sola
vez y cubre las dos columnas de sexo. Si el lector no arrastra ese valor hacia la derecha,
la mitad de las columnas queda sin grupo y se lee como si fuera un total — que es el error
que convierte «mujeres de 20 a 24» en «total país».
"""

import openpyxl
import pytest

from obsm.ingest.rem_diccionario import leer_columnas, leer_conceptos


@pytest.fixture()
def diccionario(tmp_path):
    """Reproduce la estructura verificada de una hoja P6 real."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P6"

    ws["B7"] = "REM-P6. POBLACIÓN EN CONTROL PROGRAMA DE SALUD MENTAL"
    ws["B9"] = "SECCION A.1: POBLACIÓN EN CONTROL"

    # Encabezado de tres niveles, tal como viene en el formulario.
    ws["D10"] = "T O T A L"
    ws["G10"] = "GRUPO DE EDAD (en años)"
    ws["G11"] = "0 a 4 años"
    ws["I11"] = "5 a 9 años"
    ws.merge_cells("G11:H11")      # el grupo etario cubre hombres + mujeres
    ws.merge_cells("I11:J11")
    sexos = ["Ambos sexos", "Hombres", "Mujeres",
             "Hombres", "Mujeres", "Hombres", "Mujeres"]
    for col, sexo in zip("DEFGHIJ", sexos, strict=True):
        ws[f"{col}12"] = sexo
    for i, col in enumerate("DEFGHIJ", start=1):
        ws[f"{col}13"] = f"COL{i:02d}"

    filas = [
        (14, "", "FACTORES DE RIESGO Y CONDICIONANTES", ""),
        (15, "P6221100", "VIOLENCIA", "VICTIMA"),
        (16, "P6221200", "", "AGRESOR/A"),
        (17, "P6230800", "SUICIDIO", "IDEACIÓN"),
        (18, "P6230900", "", "INTENTO"),
        (19, "P6221600", "TRASTORNOS DEL HUMOR (AFECTIVOS)", "DEPRESIÓN LEVE"),
        (20, "P6227500", "", "DEPRESIÓN MODERADA"),
    ]
    for r, codigo, b, c in filas:
        if codigo:
            ws[f"A{r}"] = codigo
        if b:
            ws[f"B{r}"] = b
        if c:
            ws[f"C{r}"] = c
        if codigo:
            for i, col in enumerate("DEFGHIJ", start=1):
                ws[f"{col}{r}"] = f"COL{i:02d}"

    ruta = tmp_path / "SP_demo.xlsx"
    wb.save(ruta)
    return ruta


class TestColumnas:
    def test_las_tres_primeras_son_el_total(self, diccionario):
        cols = {c.nombre: c for c in leer_columnas(diccionario)}
        assert cols["COL01"].es_total and cols["COL01"].sexo == "ambos"
        assert cols["COL02"].es_total and cols["COL02"].sexo == "hombres"
        assert cols["COL03"].es_total and cols["COL03"].sexo == "mujeres"

    def test_arrastra_el_grupo_etario_de_las_celdas_combinadas(self, diccionario):
        """El error que este test existe para impedir.

        `0 a 4 años` está escrito solo en G11 y la celda cubre G:H. Sin arrastrarlo,
        COL05 —mujeres de 0 a 4— quedaría sin grupo y se leería como un total.
        """
        cols = {c.nombre: c for c in leer_columnas(diccionario)}
        assert cols["COL04"].grupo_edad == "0 a 4 años"
        assert cols["COL05"].grupo_edad == "0 a 4 años"
        assert cols["COL05"].sexo == "mujeres"
        assert not cols["COL05"].es_total, "una columna de edad no puede ser un total"

    def test_cada_grupo_etario_cubre_dos_sexos(self, diccionario):
        cols = [c for c in leer_columnas(diccionario) if not c.es_total]
        por_edad: dict[str, list[str]] = {}
        for c in cols:
            por_edad.setdefault(c.grupo_edad, []).append(c.sexo)
        assert por_edad == {"0 a 4 años": ["hombres", "mujeres"],
                            "5 a 9 años": ["hombres", "mujeres"]}

    def test_la_fila_de_columnas_se_busca_y_no_se_asume(self, diccionario, tmp_path):
        # El detalle empieza en la fila 13 en 2023 pero no en todos los años. Anclar en
        # un número es la forma más común de leer mal un formulario que se reordena.
        wb = openpyxl.load_workbook(diccionario)
        ws = wb["P6"]
        ws.insert_rows(1, 5)          # se corre todo cinco filas hacia abajo
        movido = tmp_path / "movido.xlsx"
        wb.save(movido)
        assert len(leer_columnas(movido)) == len(leer_columnas(diccionario))


class TestConceptos:
    def test_extrae_codigo_y_concepto(self, diccionario):
        con = {c.codigo: c for c in leer_conceptos(diccionario)}
        assert con["P6221600"].concepto == "DEPRESIÓN LEVE"
        assert con["P6230800"].concepto == "IDEACIÓN"

    def test_arrastra_el_grupo_hacia_abajo(self, diccionario):
        """El formulario escribe el grupo una sola vez y las filas siguientes lo heredan.

        `DEPRESIÓN MODERADA` no repite «TRASTORNOS DEL HUMOR» en su fila: sin arrastre,
        ese concepto quedaría huérfano y no se podría agrupar con la depresión leve.
        """
        con = {c.codigo: c for c in leer_conceptos(diccionario)}
        assert con["P6227500"].grupo == "TRASTORNOS DEL HUMOR (AFECTIVOS)"
        assert con["P6221200"].grupo == "VIOLENCIA"
        assert con["P6230900"].grupo == "SUICIDIO"

    def test_la_etiqueta_combina_grupo_y_concepto(self, diccionario):
        con = {c.codigo: c for c in leer_conceptos(diccionario)}
        assert con["P6227500"].etiqueta == "TRASTORNOS DEL HUMOR (AFECTIVOS) · DEPRESIÓN MODERADA"

    def test_registra_que_columnas_usa_cada_concepto(self, diccionario):
        con = {c.codigo: c for c in leer_conceptos(diccionario)}
        assert con["P6221600"].columnas == ("COL01", "COL02", "COL03", "COL04",
                                            "COL05", "COL06", "COL07")

    def test_las_filas_de_encabezado_no_son_conceptos(self, diccionario):
        # «FACTORES DE RIESGO Y CONDICIONANTES» no tiene código: es un título, no un dato.
        codigos = [c.codigo for c in leer_conceptos(diccionario)]
        assert "" not in codigos
        assert len(codigos) == 6

    def test_una_hoja_que_no_existe_devuelve_vacio(self, diccionario):
        assert leer_conceptos(diccionario, hoja="P99") == []
        assert leer_columnas(diccionario, hoja="P99") == []


class TestCodigosRepetidos:
    """Un código puede aparecer dos veces en la hoja, y la segunda suele venir vacía.

    En el diccionario de 2019 las filas 118-139 repiten veintidós códigos sin grupo ni
    concepto —un listado al final del formulario—. Como el extractor guardaba en un
    diccionario, ganaba la última aparición: la ansiedad de ese año quedó etiquetada
    «Programa de rehabilitación tipo II» con el concepto en blanco, y en la serie
    publicada se veía como una caída del 99 % entre 2018 y 2020.

    El dato nunca se perdió. Se perdió su nombre, que para un observatorio es igual de
    grave: un número sin etiqueta no se puede leer ni corregir.
    """

    @pytest.fixture()
    def con_repetidos(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P6"
        ws["D10"] = "T O T A L"
        for col, sexo in zip("DEF", ["Ambos sexos", "Hombres", "Mujeres"], strict=True):
            ws[f"{col}12"] = sexo
        for i, col in enumerate("DEF", start=1):
            ws[f"{col}13"] = f"COL{i:02d}"

        # El bloque bueno: código, grupo y concepto.
        ws["A20"] = "P6232500"
        ws["B20"] = "TRASTORNOS DE ANSIEDAD"
        ws["C20"] = "TRASTORNOS DE ANSIEDAD GENERALIZADA"
        for i, col in enumerate("DEF", start=1):
            ws[f"{col}20"] = f"COL{i:02d}"

        # Más abajo, un encabezado de otra sección y el mismo código sin etiqueta.
        ws["B40"] = "PROGRAMA DE REHABILITACIÓN TIPO II"
        ws["A41"] = "P6232500"

        ruta = tmp_path / "repetidos.xlsx"
        wb.save(ruta)
        return ruta

    def test_gana_la_aparicion_que_trae_concepto(self, con_repetidos):
        con = {c.codigo: c for c in leer_conceptos(con_repetidos)}
        assert con["P6232500"].concepto == "TRASTORNOS DE ANSIEDAD GENERALIZADA"
        assert con["P6232500"].grupo == "TRASTORNOS DE ANSIEDAD"

    def test_el_codigo_aparece_una_sola_vez_en_la_salida(self, con_repetidos):
        codigos = [c.codigo for c in leer_conceptos(con_repetidos)]
        assert codigos.count("P6232500") == 1
